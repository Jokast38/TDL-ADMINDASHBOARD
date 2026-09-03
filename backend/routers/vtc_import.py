import re
import secrets
import uuid
from datetime import date, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File

from core.database import db
from core.security import hash_password, require_role
from core.utils import now_iso

router = APIRouter(prefix="/vtc-import", tags=["vtc-import"])

# Import du fichier "VTC_TAXI 2026.xlsx" fourni par l'utilisateur (un onglet
# par mois, une ligne par candidat : responsable, nom, email, centre, métier,
# n° dossier, téléphone, CMA, CPF, créneau JOUR/SOIR, notes). Ce fichier ne
# contient ni date précise de session ni formateur assigné — seulement le
# mois (nom de l'onglet) et un créneau JOUR/SOIR. Sur confirmation de
# l'utilisateur : on crée 1 session JOUR + 1 session SOIR par mois/centre/
# formation (même semaine du mois), sans formateur (à assigner ensuite
# manuellement), et on inscrit chaque candidat — payé (CPF = "OUI") ou en
# attente (avec les notes de la ligne).

_FR_MONTHS = {
    "JANVIER": 1, "FEVRIER": 2, "FÉVRIER": 2, "MARS": 3, "AVRIL": 4, "MAI": 5, "JUIN": 6,
    "JUILLET": 7, "AOUT": 8, "AOÛT": 8, "SEPTEMBRE": 9, "OCTOBRE": 10, "NOVEMBRE": 11,
    "DECEMBRE": 12, "DÉCEMBRE": 12,
}

_SHEET_RE = re.compile(r"^([A-ZÉÛÎÔÈ]+)\s+(\d{2,4})$")


def _parse_sheet_month(title: str):
    m = _SHEET_RE.match((title or "").strip().upper())
    if not m:
        return None
    month = _FR_MONTHS.get(m.group(1))
    if not month:
        return None
    year = int(m.group(2))
    if year < 100:
        year += 2000
    return year, month


def _second_week_bounds(year: int, month: int):
    """1re semaine complète (lundi-vendredi) du mois, décalée d'une semaine
    pour laisser une marge après le début du mois — approximation
    volontairement simple, ajustable ensuite depuis la page Sessions de
    stage."""
    d = date(year, month, 1)
    days_to_monday = (7 - d.weekday()) % 7
    first_monday = d + timedelta(days=days_to_monday)
    second_monday = first_monday + timedelta(days=7)
    friday = second_monday + timedelta(days=4)
    return second_monday.isoformat(), friday.isoformat()


def _match_formation(metier: str, formations: list):
    m = (metier or "").upper()
    passerelle = "PASSERELLE" in m
    if "E-LEARNING" in m or "ELEARNING" in m:
        for f in formations:
            if "en ligne" in f["title"].lower():
                return f, passerelle
        return None, passerelle
    if "TAXI" in m:
        for f in formations:
            if "taxi initiale" in f["title"].lower():
                return f, passerelle
        return None, passerelle
    if "VTC" in m:
        for f in formations:
            if f["title"].strip().lower() == "formation vtc":
                return f, passerelle
        return None, passerelle
    return None, passerelle


_CENTRE_VILLE = {
    "EPINAY-SUR-SEINE": "Épinay-sur-Seine",
    "CREIL": "Creil",
}
_CENTRE_ADRESSE = {
    "EPINAY-SUR-SEINE": "59 avenue Joffre",
    "CREIL": "",
}


@router.post("/vtc-taxi-sessions")
async def import_vtc_taxi_excel(file: UploadFile = File(...), user: dict = Depends(require_role("admin", "responsable_admission"))):
    import openpyxl
    import io as _io

    data = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(data), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fichier Excel illisible : {e}")

    formations = await db.formations.find({}, {"_id": 0, "id": 1, "title": 1, "category": 1, "price": 1}).to_list(200)

    sessions_created = 0
    sessions_existing = 0
    inscriptions_created = 0
    inscriptions_skipped = 0
    rows_skipped_no_email = 0
    unmatched_metiers = {}
    sheets_processed = []

    session_cache = {}  # (year, month, centre, formation_id, creneau) -> stage id

    for ws in wb.worksheets:
        parsed = _parse_sheet_month(ws.title)
        if not parsed:
            continue
        year, month = parsed
        sheets_processed.append(ws.title)

        header = [str(c.value or "").strip().upper() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        def col(name_part):
            for idx, h in enumerate(header):
                if name_part in h:
                    return idx
            return None

        idx_resp = col("RESPONSABLE")
        idx_prenom = col("PRENOM")
        # "NOMS" et "PRENOMS" contiennent tous deux "NOM" — on cherche l'exact
        idx_nom = next((i for i, h in enumerate(header) if h in ("NOMS", "NOM")), None)
        idx_email = col("MAIL")
        idx_centre = col("CENTRE")
        idx_metier = col("METIER")
        idx_tel = col("TELEPHONE")
        idx_cma = next((i for i, h in enumerate(header) if h == "CMA"), None)
        idx_cpf = next((i for i, h in enumerate(header) if h == "CPF"), None)
        idx_quand = col("QUAND")
        idx_notes = col("NOTES")

        for row in ws.iter_rows(min_row=2):
            values = [c.value for c in row]
            if not any(v is not None and str(v).strip() for v in values):
                continue

            def get(idx):
                if idx is None or idx >= len(values):
                    return None
                v = values[idx]
                return str(v).strip() if v is not None else None

            email = get(idx_email)
            metier = get(idx_metier)
            if not email or not metier:
                rows_skipped_no_email += 1
                continue

            formation, is_passerelle = _match_formation(metier, formations)
            if not formation:
                unmatched_metiers[metier] = unmatched_metiers.get(metier, 0) + 1
                continue

            centre_raw = (get(idx_centre) or "EPINAY-SUR-SEINE").upper()
            centre_key = centre_raw if centre_raw in _CENTRE_VILLE else "EPINAY-SUR-SEINE"
            creneau = (get(idx_quand) or "JOUR").upper()
            if creneau not in ("JOUR", "SOIR"):
                creneau = "JOUR"

            session_key = (year, month, centre_key, formation["id"], creneau)
            stage_id = session_cache.get(session_key)
            if not stage_id:
                marker = f"vtc_import_{year}-{month:02d}_{centre_key}_{formation['id']}_{creneau}"
                existing_stage = await db.stages.find_one({"import_key": marker}, {"_id": 0, "id": 1})
                if existing_stage:
                    stage_id = existing_stage["id"]
                    sessions_existing += 1
                else:
                    date_debut, date_fin = _second_week_bounds(year, month)
                    stage_id = str(uuid.uuid4())
                    label = "Passerelle" if is_passerelle else formation["title"]
                    await db.stages.insert_one({
                        "id": stage_id, "formation_id": formation["id"], "formation_titre": formation["title"],
                        "date_debut": date_debut, "date_fin": date_fin,
                        "lieu_adresse": _CENTRE_ADRESSE.get(centre_key, ""), "lieu_ville": _CENTRE_VILLE[centre_key],
                        "capacite_max": 25, "animateur_id": None, "statut": "planifie", "nb_inscrits": 0,
                        "notes": f"Créneau : {creneau}. {label} — importé depuis le fichier Excel VTC_TAXI 2026 ({ws.title}). Formateur à assigner.",
                        "import_key": marker, "created_at": now_iso(),
                    })
                    sessions_created += 1
                session_cache[session_key] = stage_id

            existing_insc = await db.inscriptions.find_one(
                {"student_email": email.lower(), "formation_id": formation["id"], "source": "excel_import_vtc_taxi_2026"},
                {"_id": 0, "id": 1},
            )
            if existing_insc:
                inscriptions_skipped += 1
                continue

            name = f"{get(idx_prenom) or ''} {get(idx_nom) or ''}".strip() or email.split("@")[0]
            existing_user = await db.users.find_one({"email": email.lower()})
            if existing_user:
                user_id = existing_user["id"]
            else:
                user_id = str(uuid.uuid4())
                await db.users.insert_one({
                    "id": user_id, "email": email.lower(), "name": name, "role": "etudiant",
                    "phone": get(idx_tel), "password_hash": hash_password(secrets.token_urlsafe(12)),
                    "created_at": now_iso(), "active": True,
                })

            cpf_ok = (get(idx_cpf) or "").upper() == "OUI"
            note_parts = []
            if get(idx_notes):
                note_parts.append(get(idx_notes))
            if get(idx_resp):
                note_parts.append(f"Chargé(e) : {get(idx_resp)}")
            if get(idx_cma):
                note_parts.append(f"CMA : {get(idx_cma)}")
            note_parts.append(f"Créneau : {creneau} — importé du fichier Excel VTC_TAXI 2026 ({ws.title})")

            insc_id = str(uuid.uuid4())
            inscription = {
                "id": insc_id, "formation_id": formation["id"], "formation_title": formation["title"],
                "category": formation["category"], "student_id": user_id, "student_name": name,
                "student_email": email.lower(), "student_phone": get(idx_tel),
                "price": formation.get("price", 0),
                # CPF = "OUI" dans le fichier -> dossier de financement CPF déjà
                # validé (statut dédié "cpf_valide", distinct d'un règlement
                # direct "paid") ; sinon en attente de validation CPF.
                "payment_status": "cpf_valide" if cpf_ok else "cpf_attente",
                "status": "active", "contact_status": "finalisee" if cpf_ok else "a_contacter",
                "notes": " · ".join(note_parts), "created_at": now_iso(),
                "source": "excel_import_vtc_taxi_2026", "session": stage_id, "stage_id": stage_id,
                "center": _CENTRE_VILLE[centre_key],
            }
            if cpf_ok:
                inscription["paid_at"] = now_iso()
                inscription["amount_paid"] = formation.get("price", 0)
            await db.inscriptions.insert_one(inscription)

            dossier_id = str(uuid.uuid4())
            await db.dossiers.insert_one({
                "id": dossier_id, "inscription_id": insc_id, "student_id": user_id,
                "formation_id": formation["id"], "formation_title": formation["title"],
                "category": formation["category"], "student_name": name, "student_email": email.lower(),
                "status": "nouveau", "notes": "", "assigned_to": None,
                "documents_requis": formation.get("documents_requis", []),
                "trello_card_id": None, "trello_card_url": None,
                "documents": [], "created_at": now_iso(), "updated_at": now_iso(),
            })
            inscriptions_created += 1

    if not sheets_processed:
        raise HTTPException(status_code=400, detail="Aucun onglet mensuel reconnu (ex: \"SEPTEMBRE 26\") dans ce fichier")

    return {
        "sheets_processed": sheets_processed,
        "sessions_created": sessions_created,
        "sessions_already_existing": sessions_existing,
        "inscriptions_created": inscriptions_created,
        "inscriptions_already_existing": inscriptions_skipped,
        "rows_skipped_missing_email_or_metier": rows_skipped_no_email,
        "unmatched_metiers": unmatched_metiers,
    }
