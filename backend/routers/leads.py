import re
import uuid
from typing import Dict, List, Optional, Any
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File

from core.database import db
from core.security import require_role
from core.utils import now_iso
from core.config import ROLES_LEADS
from models.lead import LeadIn, LeadUpdate, LeadImportJsonIn, LeadRelanceIn, LeadRelanceSingleIn
from services.email import send_email

router = APIRouter(prefix="/leads", tags=["leads"])


# ---- Helpers normalisation ----

def _strip_accents(s: str) -> str:
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def _clean_header(h) -> str:
    if not h:
        return ""
    return _strip_accents(str(h).strip()).lower()


def _clean_phone(raw) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = s.replace(" ", "").replace("-", "").replace(".", "")
    if not s or not any(c.isdigit() for c in s):
        return None
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) == 9:
        s = "0" + digits
    elif len(digits) == 10:
        s = digits
    return s if len(s) >= 8 else None


def _find_field(raw: dict, keywords: list, exclude_keywords: Optional[list] = None) -> str:
    for k, v in raw.items():
        if not k or v in (None, ""):
            continue
        kk = _clean_header(k)
        if exclude_keywords and any(ex in kk for ex in exclude_keywords):
            continue
        if any(kw in kk for kw in keywords):
            return str(v).strip()
    return ""


def _normalize_interest(raw: str) -> str:
    if not raw:
        return ""
    sl = _strip_accents(raw.strip()).lower()
    has = lambda kw: kw in sl
    if has("passerelle") and has("vtc"):
        return "Passerelle VTC"
    if has("passerelle") and has("taxi"):
        return "Passerelle Taxi"
    if has("passerelle"):
        return "Passerelle"
    if has("mobilit") and has("taxi"):
        return "Mobilité Taxi"
    if has("mobilit"):
        return "Mobilité"
    if has("vtc"):
        return "VTC"
    if has("taxi"):
        return "Taxi"
    if has("caces"):
        return "CACES"
    if has("ssiap"):
        return "SSIAP"
    if has("permis b"):
        return "Permis B"
    if has("crm"):
        return "CRM"
    if has("stage"):
        return "Stage"
    return raw.strip()


def _detect_tdl_planning_columns(headers: list) -> Optional[dict]:
    clean = [_clean_header(h) for h in headers]
    has_responsable = any("responsable" in c or "inscription" in c or "provenance" in c for c in clean)
    has_metier = any("metier" in c or "formation" in c or "colonne" in c for c in clean[:2])
    if not (has_responsable and has_metier):
        return None

    def find(*kws) -> Optional[int]:
        for i, c in enumerate(clean):
            if any(kw in c for kw in kws):
                return i
        return None

    interet_idx = find("interet", "formation", "metier", "colonne")
    if interet_idx is None:
        interet_idx = 0

    prenom_idx = find("prenom", "firstname", "first name")
    nom_idx = find("nom", "lastname", "last name")
    if prenom_idx is None and nom_idx is None:
        prenom_idx, nom_idx = 3, 4
    elif prenom_idx is None:
        prenom_idx = 3 if nom_idx != 3 else 4
    elif nom_idx is None:
        nom_idx = 4 if prenom_idx != 4 else 3

    email_idx = find("mail", "email", "courriel", "@")
    if email_idx is None:
        email_idx = 5

    phone_idx = find("tel", "phone", "portable", "mobile", "gsm", "numero")
    if phone_idx is None:
        phone_idx = 6

    comment_idx = find("commentaire", "comment", "note")
    if comment_idx is None:
        comment_idx = min(10, len(clean) - 1) if clean else 10

    return {
        "interet": interet_idx,
        "responsable": find("responsable", "inscription", "provenance") or 1,
        "centre": find("ville", "centre") or 2,
        "prenom": prenom_idx, "nom": nom_idx,
        "email": email_idx, "phone": phone_idx, "commentaire": comment_idx,
    }


def _normalize_lead_tdl(row: tuple, col_map: dict, source: str) -> Optional[dict]:
    def get(idx) -> str:
        if idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    prenom = get(col_map["prenom"])
    nom = get(col_map["nom"])
    name = (prenom + " " + nom).strip()
    email_raw = get(col_map["email"])
    email = email_raw.split("/")[0].split(",")[0].strip().lower() or None
    phone = _clean_phone(row[col_map["phone"]] if col_map["phone"] < len(row) else None)
    interest = _normalize_interest(get(col_map["interet"]))
    notes = get(col_map["commentaire"])
    if not name and not email and not phone:
        return None
    tags = ["a_appeler"] if phone and not email else []
    return {
        "id": str(uuid.uuid4()), "name": name or email or phone or "Lead sans nom",
        "email": email, "phone": phone, "interest": interest, "notes": notes,
        "tags": tags, "contacted": False, "status": "nouveau",
        "source": source, "created_at": now_iso(), "updated_at": now_iso(),
    }


def _normalize_lead(raw: dict) -> Optional[dict]:
    prenom = _find_field(raw, ["prenom", "first name", "firstname"])
    nom = _find_field(raw, ["nom", "name", "lastname", "last name"], exclude_keywords=["prenom", "mail"])
    full_name = (prenom + " " + nom).strip()
    email = ""
    for k, v in raw.items():
        if v in (None, ""):
            continue
        kk = _clean_header(str(k))
        if kk in ("@", "mail", "email", "courriel") or "mail" in kk or kk == "@":
            email = str(v).strip().split("/")[0].split(",")[0].strip().lower()
            break
    email = email or None
    phone = _clean_phone(_find_field(raw, ["tel", "phone", "portable", "mobile", "gsm", "numero"])) or None
    if not email and not phone:
        return None
    interest = _normalize_interest(_find_field(raw, ["interet", "formation", "metier"]))
    name = full_name or email or phone
    tags = ["a_appeler"] if phone and not email else []
    return {
        "id": str(uuid.uuid4()), "name": name, "email": email, "phone": phone,
        "interest": interest, "notes": "", "tags": tags, "contacted": False,
        "status": "nouveau", "source": raw.get("_source", "import"),
        "created_at": now_iso(), "updated_at": now_iso(),
    }


def _looks_like_header(row: tuple) -> bool:
    """True si la ligne ressemble à des intitulés de colonnes plutôt qu'à des données réelles."""
    for cell in row:
        if cell is None:
            continue
        s = str(cell).strip()
        if not s:
            continue
        if '@' in s and '.' in s.split('@')[-1]:
            return False
        if len(re.sub(r'\D', '', s)) >= 8:
            return False
    return True


_INTEREST_KEYWORDS_SET = frozenset((
    "vtc", "taxi", "passerelle", "mobilit", "caces", "ssiap",
    "permis", "crm", "stage", "formation",
))


def _col_looks_like_interest(data_rows: list, col_idx: int) -> bool:
    """Vérifie si une colonne contient des mots-clés de formation (VTC, Taxi…)
    et non des prénoms d'agents (Anaïs, vanessa…)."""
    for row in data_rows[:20]:
        v = row[col_idx] if col_idx < len(row) else None
        if not v:
            continue
        sl = _strip_accents(str(v).strip()).lower()
        if any(kw in sl for kw in _INTEREST_KEYWORDS_SET):
            return True
    return False


def _infer_columns_by_content(data_rows: list) -> dict:
    """Déduit les positions de colonnes à partir du contenu.

    Stratégie : une fois l'email trouvé (colonne avec @), les deux colonnes
    qui le précèdent sont NOM et PRÉNOM. La colonne 0 n'est traitée comme
    intérêt que si elle contient des mots-clés de formation — sinon c'est
    probablement le prénom de l'agent responsable (ex. Anaïs, vanessa…).
    """
    n_cols = max((len(r) for r in data_rows), default=0)
    email_col = phone_col = None

    for col_idx in range(n_cols):
        email_hits = phone_hits = 0
        for row in data_rows[:30]:
            v = row[col_idx] if col_idx < len(row) else None
            if v is None:
                continue
            s = str(v).strip()
            if '@' in s:
                email_hits += 1
            elif len(re.sub(r'\D', '', s)) >= 8:
                phone_hits += 1
        if email_hits > 0 and email_col is None:
            email_col = col_idx
        if phone_hits > 0 and phone_col is None:
            phone_col = col_idx

    # Ancre : colonne email (ou téléphone si aucun email)
    anchor = email_col if email_col is not None else phone_col

    # Les deux colonnes immédiatement avant l'ancre → nom, prénom
    nom_col = (anchor - 1) if anchor is not None and anchor >= 1 else None
    prenom_col = (anchor - 2) if anchor is not None and anchor >= 2 else None

    # Colonne 0 = intérêt uniquement si elle contient des mots-clés de formation
    # (si c'est "Anaïs" ou "vanessa" → pas un intérêt, on laisse None)
    interet_col = None
    if prenom_col != 0 and nom_col != 0 and _col_looks_like_interest(data_rows, 0):
        interet_col = 0

    # Commentaire : dernière colonne peuplée après le téléphone
    comment_col = None
    if phone_col is not None:
        for c in range(n_cols - 1, phone_col, -1):
            if any(c < len(r) and r[c] not in (None, "") for r in data_rows[:20]):
                comment_col = c
                break

    return {
        "interet": interet_col,
        "prenom": prenom_col,
        "name": nom_col,
        "email": email_col,
        "phone": phone_col,
        "commentaire": comment_col,
    }


def _normalize_lead_from_colmap(row: tuple, col_map: dict, source: str) -> Optional[dict]:
    """Normalise un lead depuis un fichier sans en-têtes via les colonnes détectées par contenu."""
    def get(idx) -> str:
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    email_raw = get(col_map.get("email"))
    email = (email_raw.split("/")[0].split(",")[0].strip().lower() or None) if email_raw else None
    phone = _clean_phone(
        row[col_map["phone"]] if col_map.get("phone") is not None and col_map["phone"] < len(row) else None
    )
    if not email and not phone:
        return None
    prenom = get(col_map.get("prenom"))
    nom = get(col_map.get("name"))
    name = (prenom + " " + nom).strip()
    interest = _normalize_interest(get(col_map.get("interet")))
    notes = get(col_map.get("commentaire"))
    tags = ["a_appeler"] if phone and not email else []
    return {
        "id": str(uuid.uuid4()),
        "name": name or email or phone,
        "email": email, "phone": phone,
        "interest": interest, "notes": notes,
        "tags": tags, "contacted": False,
        "status": "nouveau", "source": source,
        "created_at": now_iso(), "updated_at": now_iso(),
    }


async def _insert_leads_dedup(leads: list) -> dict:
    inserted, skipped = 0, 0
    for lead in leads:
        existing = None
        if lead.get("email"):
            existing = await db.leads.find_one({"email": lead["email"]})
        if not existing and lead.get("phone"):
            existing = await db.leads.find_one({"phone": lead["phone"]})
        if existing:
            skipped += 1
            continue
        await db.leads.insert_one(lead)
        inserted += 1
    return {"inserted": inserted, "skipped_duplicates": skipped}


# ---- Routes ----

@router.get("")
async def list_leads(
    tag: Optional[str] = None,
    status: Optional[str] = None,
    contacted: Optional[bool] = None,
    has_email: Optional[bool] = None,
    has_phone: Optional[bool] = None,
    q: Optional[str] = None,
    user: dict = Depends(require_role(*ROLES_LEADS))
):
    query: Dict[str, Any] = {}
    if tag: query["tags"] = tag
    if status: query["status"] = status
    if contacted is not None: query["contacted"] = contacted
    if has_email is not None:
        query["email"] = {"$ne": None} if has_email else None
    if has_phone is not None:
        query["phone"] = {"$ne": None} if has_phone else None
    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"email": {"$regex": q, "$options": "i"}},
            {"phone": {"$regex": q, "$options": "i"}},
        ]
    return await db.leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)


@router.post("")
async def create_lead(payload: LeadIn, user: dict = Depends(require_role(*ROLES_LEADS))):
    lead = payload.model_dump()
    lead["id"] = str(uuid.uuid4())
    if lead.get("phone") and not lead.get("email") and "a_appeler" not in lead["tags"]:
        lead["tags"].append("a_appeler")
    lead["contacted"] = False
    lead["status"] = "nouveau"
    lead["source"] = "manuel"
    lead["created_at"] = now_iso()
    lead["updated_at"] = now_iso()
    await db.leads.insert_one(lead)
    lead.pop("_id", None)
    return lead


@router.put("/{lid}")
async def update_lead(lid: str, payload: LeadUpdate, user: dict = Depends(require_role(*ROLES_LEADS))):
    existing = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Lead introuvable")
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    update["updated_at"] = now_iso()
    await db.leads.update_one({"id": lid}, {"$set": update})
    return await db.leads.find_one({"id": lid}, {"_id": 0})


@router.delete("/{lid}")
async def delete_lead(lid: str, user: dict = Depends(require_role(*ROLES_LEADS))):
    await db.leads.delete_one({"id": lid})
    return {"ok": True}


@router.post("/bulk-delete")
async def bulk_delete_leads(payload: Dict[str, List[str]], user: dict = Depends(require_role(*ROLES_LEADS))):
    ids = payload.get("lead_ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="Aucun lead sélectionné")
    result = await db.leads.delete_many({"id": {"$in": ids}})
    return {"deleted": result.deleted_count}


@router.post("/import/json")
async def import_leads_json(payload: LeadImportJsonIn, user: dict = Depends(require_role(*ROLES_LEADS))):
    leads = [l for raw in payload.leads if (l := _normalize_lead({**raw, "_source": "import_json"}))]
    return await _insert_leads_dedup(leads)


@router.post("/import/xlsx")
async def import_leads_xlsx(file: UploadFile = File(...), user: dict = Depends(require_role(*ROLES_LEADS))):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl n'est pas installé (pip install openpyxl)")
    data = await file.read()
    try:
        wb = openpyxl.load_workbook(__import__("io").BytesIO(data), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fichier Excel illisible: {e}")

    all_leads = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Première ligne non-vide (au moins 2 cellules remplies)
        first_idx = next(
            (i for i, row in enumerate(rows) if sum(1 for c in row if c is not None) >= 2),
            None
        )
        if first_idx is None:
            continue

        candidate = rows[first_idx]
        source = f"import_xlsx_{sheet_name}"

        # 1. Format TDL planning ?
        col_map = _detect_tdl_planning_columns(list(candidate))
        if col_map:
            for row in rows[first_idx + 1:]:
                if not any(row):
                    continue
                lead = _normalize_lead_tdl(row, col_map, f"import_xlsx_tdl_{sheet_name}")
                if lead:
                    all_leads.append(lead)
            continue

        # 2. La première ligne ressemble-t-elle à des en-têtes ?
        if _looks_like_header(candidate):
            clean_headers = [_clean_header(h) for h in candidate]
            if clean_headers and not clean_headers[0]:
                clean_headers[0] = "prenom"
            for row in rows[first_idx + 1:]:
                if not any(row):
                    continue
                raw = {clean_headers[i]: row[i] for i in range(min(len(clean_headers), len(row))) if clean_headers[i]}
                raw["_source"] = source
                lead = _normalize_lead(raw)
                if lead:
                    all_leads.append(lead)
        else:
            # 3. Pas d'en-têtes : détecter les colonnes par le contenu
            data_rows = [r for r in rows[first_idx:] if any(r)]
            col_map_content = _infer_columns_by_content(data_rows)
            for row in data_rows:
                lead = _normalize_lead_from_colmap(row, col_map_content, source)
                if lead:
                    all_leads.append(lead)

    if not all_leads:
        raise HTTPException(status_code=400, detail="Aucune donnée exploitable trouvée dans le fichier")
    return await _insert_leads_dedup(all_leads)


@router.post("/import/from-inscriptions")
async def import_leads_from_inscriptions(user: dict = Depends(require_role(*ROLES_LEADS))):
    inscriptions = await db.inscriptions.find({}, {"_id": 0}).to_list(5000)
    leads = [
        l for ins in inscriptions
        if (l := _normalize_lead({
            "name": ins.get("student_name"), "email": ins.get("student_email"),
            "phone": ins.get("student_phone"), "interest": ins.get("formation_title", ""),
            "_source": "inscription",
        }))
    ]
    return await _insert_leads_dedup(leads)


@router.get("/{lid}/emails")
async def list_lead_emails(lid: str, user: dict = Depends(require_role(*ROLES_LEADS))):
    return await db.email_logs.find(
        {"lead_id": lid}, {"_id": 0, "body": 0}
    ).sort("created_at", -1).to_list(200)


@router.post("/relance/single")
async def send_relance_single(payload: LeadRelanceSingleIn, user: dict = Depends(require_role(*ROLES_LEADS))):
    lead = await db.leads.find_one({"id": payload.lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead introuvable")
    if not lead.get("email"):
        return {"sent": False, "reason": "no_email"}
    body = payload.body.replace("{{name}}", lead.get("name", ""))
    log = await send_email(lead["email"], payload.subject, body, extra={"lead_id": payload.lead_id})
    if log["status"] not in ("sent", "mocked"):
        return {"sent": False, "reason": log["status"]}
    update = {"updated_at": now_iso()}
    if payload.mark_contacted:
        update["contacted"] = True
        update["last_contacted_at"] = now_iso()
        update["status"] = "contacte"
    await db.leads.update_one({"id": payload.lead_id}, {"$set": update})
    if payload.add_tag:
        await db.leads.update_one({"id": payload.lead_id}, {"$addToSet": {"tags": payload.add_tag}})
    return {"sent": True}


@router.post("/relance")
async def send_relance(payload: LeadRelanceIn, user: dict = Depends(require_role(*ROLES_LEADS))):
    if not payload.lead_ids:
        raise HTTPException(status_code=400, detail="Aucun lead sélectionné")
    leads = await db.leads.find({"id": {"$in": payload.lead_ids}}, {"_id": 0}).to_list(len(payload.lead_ids))
    sent, no_email = 0, 0
    for lead in leads:
        if not lead.get("email"):
            no_email += 1
            continue
        body = payload.body.replace("{{name}}", lead.get("name", ""))
        log = await send_email(lead["email"], payload.subject, body, extra={"lead_id": lead["id"]})
        if log["status"] not in ("sent", "mocked"):
            continue
        sent += 1
        update = {"updated_at": now_iso()}
        if payload.mark_contacted:
            update["contacted"] = True
            update["last_contacted_at"] = now_iso()
            update["status"] = "contacte"
        await db.leads.update_one({"id": lead["id"]}, {"$set": update})
        if payload.add_tag:
            await db.leads.update_one({"id": lead["id"]}, {"$addToSet": {"tags": payload.add_tag}})
    return {"sent": sent, "skipped_no_email": no_email}
